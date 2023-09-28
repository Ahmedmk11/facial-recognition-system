# ---------------------------------------
# Imports & Initialization
# ---------------------------------------

from datetime import datetime
import json
import os
from typing import *
from flask import Flask, make_response, redirect, request, jsonify, send_file, session, send_from_directory, url_for
from flask_cors import CORS
from mysql.connector.connection import MySQLConnection
import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import Alignment
import pandas as pd
import base64
from facial_recognition.count_faces import count_faces
from facial_recognition.find_matches import find_matches
import io
from PIL import Image

app = Flask(__name__, static_folder='../dist')
CORS(app, origins="http://127.0.0.1:5173", supports_credentials=True)

# ---------------------------------------
# Database
# ---------------------------------------

def resize_image(base64_string):
    # Remove the 'data:image/jpeg;base64,' part from the base64 string
    base64_string = base64_string.split(',')[1]

    # Decode the base64 string to bytes
    image_data = base64.b64decode(base64_string)

    # Open the image using Pillow
    image = Image.open(io.BytesIO(image_data))

    # Calculate dimensions for cropping the image to a 1:1 aspect ratio
    width, height = image.size
    new_size = min(width, height)
    left = (width - new_size) / 2
    top = (height - new_size) / 2
    right = (width + new_size) / 2
    bottom = (height + new_size) / 2

    # Crop the image to a 1:1 aspect ratio
    cropped_image = image.crop((left, top, right, bottom))

    # Resize the cropped image to 150x150 pixels
    resized_image = cropped_image.resize((150, 150), Image.ANTIALIAS)

    # Convert the resized image back to bytes
    buffer = io.BytesIO()
    resized_image.save(buffer, format="JPEG")
    resized_image_data = buffer.getvalue()

    # Encode the resized image as base64 and add the URI part back
    resized_base64 = "data:image/jpeg;base64," + base64.b64encode(resized_image_data).decode("utf-8")

    return resized_base64

def base64_to_blob(base64_data):
    parts = base64_data.split(',')
    if len(parts) != 2:
        raise ValueError("Invalid Base64 data format")
    content_type, data = parts
    content_type = content_type.split(';')[0]
    decoded_data = base64.b64decode(data)

    return decoded_data

# ---------------------------------------
# Database
# ---------------------------------------

def create_db_connection() -> MySQLConnection:
    connection = MySQLConnection(
        host='localhost',
        user='root',
        password= os.environ.get('DATABASE_PASSWORD'),
        database='xceed'
    )
    return connection

def call_procedure(procedure_name: str, *params: str):
    conn = create_db_connection()
    cursor = conn.cursor()
    cursor.callproc(procedure_name, params)
    results = cursor.stored_results()
    
    result_sets = []
    for result in results:
        result_sets.append(result.fetchall())

    cursor.close()
    conn.close()

    if result_sets and result_sets[0] and result_sets[0][0] == ('OK',):
        return 'OK'
    
    return result_sets

def update_user_in_db(userId, newFirstname, newLastname, newEmail, newUsername, newJobtitle, newStreetAddress,
                      newLocation, newPhoneNumber, newRole, newBirthdate, newEmploymentStatus, newDepartmentId, newImage, debugMode):
    try:
        resized = resize_image(newImage)
        newImageBlob = base64_to_blob(resized)
        conn = create_db_connection()
        cursor = conn.cursor()
        cursor.callproc("UpdateUser", (userId, newFirstname, newLastname, newEmail, newUsername, newJobtitle,
                                       newStreetAddress, newLocation, newPhoneNumber, newRole, newBirthdate,
                                       newEmploymentStatus, newDepartmentId, newImageBlob, debugMode))
        conn.commit()
        cursor.close()
        conn.close()
        return 'OK'
    except Exception as e:
        return str(e)
    
def add_attendance(userId):
    try:
        conn = create_db_connection()
        cursor = conn.cursor()
        cursor.callproc("InsertAttendance", (userId,))
        conn.commit()
        cursor.close()
        conn.close()
        return 'OK'
    except Exception as e:
        return str(e)

def get_data(query: str) -> Dict[str, Any]:
    connection = create_db_connection()
    cursor = connection.cursor(dictionary=True)

    cursor.execute(query)

    data = cursor.fetchall()

    cursor.close()
    connection.close()

    return data

def call_insert_procedure(
    procedure_name: str,
    *params: str
):
    try:
        connection = create_db_connection()
        cursor = connection.cursor()
        cursor.callproc(procedure_name, params)
        connection.commit()
        return

    except mysql.connector.Error as err:
        print(f"Error: {err}")
        connection.rollback()
        return -1
    finally:
        cursor.close()
        connection.close()

# ---------------------------------------
# Session
# ---------------------------------------

app.config['SESSION_TYPE'] = 'cookie'
app.config['SESSION_PERMANENT'] = True
app.config['SESSION_USE_SIGNER'] = True
app.config['SESSION_KEY_PREFIX'] = 'xceed_'
app.config['SESSION_COOKIE_SAMESITE'] = 'None'
app.config['SESSION_COOKIE_SECURE'] = True

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')
api_key = os.environ.get('API_KEY')
debug_mode = os.environ.get('DEBUG')

if debug_mode and debug_mode.lower() == 'true':
    app.debug = True

# ---------------------------------------
# POST requests
# ---------------------------------------

@app.route('/register', methods=['POST'])
def register():
    if request.is_json:
        data = request.get_json()
        firstname = data.get('firstname')
        lastname = data.get('lastname')
        email = data.get('email')
        username = data.get('username')
        street_address = data.get('street_address')
        location = data.get('location')
        phone_number = data.get('phone_number')
        screenshot = data.get('screenshot')
        screenshot = resize_image(screenshot)
        screenshotBlob = base64_to_blob(screenshot)

        date_format = '%Y-%m-%dT%H:%M:%S.%fZ'
        parsed_date = datetime.strptime(data.get('birthdate'), date_format)
        birthdate = parsed_date.date()

        if not (firstname and lastname and username and email and street_address and location and phone_number and birthdate and screenshot):
            error_response = {'error': 'Invalid or missing data'}
            return jsonify(error_response), 400

        procedure_name = "InsertUser"
        params = (firstname, lastname, email, username, street_address, location, phone_number, birthdate, screenshotBlob)
        call_insert_procedure(procedure_name, *params)
        res = call_procedure('GetIDByUserName', username)
        inserted_id = res[0][0][0]

        if inserted_id != -1:
            response_data = {'message': 'Registration successful'}
            return jsonify(response_data), 200
        else:
            error_response = {'error': 'Failed to insert data'}
            return jsonify(error_response), 500  # You can use a more appropriate HTTP status code

    else:
        error_response = {'error': 'Invalid request format'}
        return jsonify(error_response), 400

@app.route('/api/facial-recognition', methods=['POST'])
def get_webcam_frames():
    frame_data = request.json.get('frameData')
    un = request.json.get('un')

    data = call_procedure('GetUserPicture', un)
    bin_string = data[0][0][0]
    known_image_base64 = [base64.b64encode(bin_string).decode('utf-8')]

    if frame_data:
        frame_data = resize_image(frame_data)
        test_image_base64 = frame_data.replace('data:image/jpeg;base64,', '')
        faces_num = count_faces(test_image_base64)
        if faces_num > 1:
            return jsonify({'message': faces_num}), 403
        elif faces_num == 0:
            return jsonify({'message': faces_num}), 403
        elif faces_num == 1:
            isMatch = find_matches(test_image_base64, known_image_base64)
            if isMatch:
                queryResult = call_procedure('GetAllUsersByUsername', un)
                session['user'] = {
                    'user_id' : queryResult[0][0][0],
                    'username' : queryResult[0][0][4],
                    'email' : queryResult[0][0][3],
                    'firstname' : queryResult[0][0][1],
                    'lastname' : queryResult[0][0][2],
                }
                uid = queryResult[0][0][0]
                insertResult = add_attendance(uid)
                if (insertResult != 'OK'):
                    return jsonify({'message': 'Login Failed, Cannot insert attendance'}), 500
                else:
                    return jsonify({'message': 'Login Successful'}), 200
            else:
                allPics = call_procedure('GetAllUsersPictures')
                [allIDs] = call_procedure('GetAllUserIDs')
                bin_strings = allPics[0]
                print('bin_strings', bin_strings)
                index = 0
                for bin_tuple in bin_strings:
                    bin_s = bin_tuple[0]
                    possible_image_base64 = [base64.b64encode(bin_s).decode('utf-8')]
                    isMatchAnotherUsername = find_matches(test_image_base64, possible_image_base64)
                    fid = allIDs[index]
                    print('fid', fid)
                    if isMatchAnotherUsername:
                        queryResult = call_procedure('GetAllUsersByUsername', un)
                        [[found_user]] = call_procedure('GetUserByID', fid[0])
                        if queryResult and found_user:
                            fn = f'{queryResult[0][0][1]} {queryResult[0][0][2]}'
                            eid = queryResult[0][0][0]
                            user_in_picture_name = f'{found_user[1]} {found_user[2]}'
                            user_in_picture_id = found_user[0]

                            cnx = create_db_connection()
                            cursor = cnx.cursor()
                            cursor.callproc('InsertNotification', args=(user_in_picture_id, eid, user_in_picture_name, fn))
                            cnx.commit()
                            cursor.close()
                            cnx.close()

                            return jsonify({'message': 'Violaton happened', 'violation': True, 'vname': user_in_picture_name}), 403
                        else:
                            return jsonify({'message': 'Cant Find user'}), 404
                    else:
                        index += 1
                        continue
                return jsonify({'message': 'Not recognized'}), 418

@app.route('/login', methods=['POST'])
def login():
    if request.is_json:
        data = request.get_json()
        username = data.get('username')
    if not (username):
        error_response = {'error': 'Invalid or missing data'}
        return jsonify(error_response), 400
    queryResult = call_procedure('GetAllUsersByUsername', username)

    if not queryResult:
        return jsonify({'message': f'We couldn\'t find an account with username: {username}'}), 404
    else:
        return jsonify({'message': 'Username Found'}), 200

@app.route('/clear-session-user', methods=['POST'])
def clear_session_user():
    try:
        session.clear()
        return jsonify({'message': 'Session cleared successfully'}), 200
    except:
        return jsonify({'message': 'Error clearing session'}), 500
    
@app.route('/api/update-user', methods=['POST'])
def update_user():
    data = request.json
    userId = data.get('userId')
    newFirstname = data.get('newFirstname')
    newLastname = data.get('newLastname')
    newEmail = data.get('newEmail')
    newUsername = data.get('newUsername')
    newJobtitle = data.get('newJobtitle')
    newStreetAddress = data.get('newStreetAddress')
    newLocation = data.get('newLocation')
    newPhoneNumber = data.get('newPhoneNumber')
    newRole = data.get('newRole')
    newBirthdate = data.get('newBirthdate')
    newEmploymentStatus = data.get('newEmploymentStatus')
    newDepartmentId = data.get('newDepartmentId')
    newImage = data.get('newImage')
    debugMode = data.get('debugMode')

    result = update_user_in_db(userId, newFirstname, newLastname, newEmail, newUsername, newJobtitle, newStreetAddress,
                               newLocation, newPhoneNumber, newRole, newBirthdate, newEmploymentStatus, newDepartmentId, newImage, debugMode)

    return jsonify({'result': result})
    
# ---------------------------------------
# GET requests
# ---------------------------------------

@app.route('/api/check-session', methods=['GET'])
def check_session():
    if 'user' in session:
        return jsonify({'authenticated': True})
    else:
        return jsonify({'authenticated': False})
    
@app.route('/api/check-email', methods=['GET'])
def check_email():
    email = request.args.get('email')
    
    if not email:
        return jsonify({'emailInUse': False})
    
    count = call_procedure('GetEmailCount' , email)
    if count[0][0][0] != 0:
        return jsonify({'emailInUse': True})
    
    return jsonify({'emailInUse': False})

@app.route('/api/check-username', methods=['GET'])
def check_username():
    username = request.args.get('username')
    
    if not username:
        return jsonify({'usernameInUse': False})

    count = call_procedure('GetUsernameCount', username)

    if count[0][0][0] != 0:
        return jsonify({'usernameInUse': True})
    
    return jsonify({'usernameInUse': False})

@app.route('/api/check-number', methods=['GET'])
def check_number():
    number = request.args.get('number')
    
    if not number:
        return jsonify({'numberInUse': False})
    
    count = call_procedure('GetPhoneNumberCount', number)
    if count[0][0][0] != 0:
        return jsonify({'numberInUse': True})
    
    return jsonify({'numberInUse': False})

@app.route('/api/check-username-exists', methods=['GET'])
def check_username_exists():
    username = request.args.get('username')
    
    if not username:
        return jsonify({'username_exists': False})

    count = call_procedure('GetUsernameCount' ,username)
    if count != 0:
        return jsonify({'username_exists': True})
    
    return jsonify({'username_exists': False})

@app.route('/api/check-role', methods=['GET'])
def check_role():
    uid = session['user']['user_id']
    if not uid:
        return jsonify({'error': 'No user found'})
    
    role = call_procedure('GetUserRole' ,uid)
    if not role:
        return jsonify({'error': 'No role found'})
    
    return jsonify({'role': role[0][0][0]})

@app.route('/api/check-name-site', methods=['GET'])
def check_name_site():
    uid = request.args.get('uid')
    if uid == '':
        uid = session['user']['user_id']
    if not uid:
        return jsonify({'error': 'No user found'})
    
    [[nameAndSite]] = call_procedure('GetUserDepartmentAndSite' ,uid)
    return jsonify({'nameAndSite': nameAndSite})

@app.route('/api/get-all-users', methods=['GET'])
def get_users():
    [users] = call_procedure('GetAllUsers')
    return jsonify({'users': users})

@app.route('/api/get-all-departments', methods=['GET'])
def get_departments():
    [departments] = call_procedure('GetAllDepartments')
    return jsonify({'departments': departments})

@app.route('/api/get-user-attendance', methods=['GET'])
def get_user_attendance():
    uid = request.args.get('uid')
    if not uid:
        return jsonify({'error': 'No user found'})
    
    [attendance] = call_procedure('GetUserAttendance', uid)
    return jsonify({'attendance': attendance})

@app.route('/api/get-user-id', methods=['GET'])
def get_user_by_id():
    uid = request.args.get('uid')
    if uid == '':
        uid = session['user']['user_id']
    
    if not uid:
        return jsonify({'error': 'No user found'})
    
    [[user]] = call_procedure('GetUserByID', uid)
    user_list = list(user)
    
    base64_string = 'data:image/jpeg;base64,' + base64.b64encode(user_list[13]).decode('utf-8')
    resized = resize_image(base64_string)
    resized = resized.split(',')[1]
    print(resize_image)
    user_list[13] = resized
    user = tuple(user_list)
    return jsonify({'user': user})

@app.route('/api/get-dep-id', methods=['GET'])
def get_dep_id():
    name = request.args.get('name')
    site = request.args.get('site')
    if not name or not site:
        return jsonify({'error': 'No department found'})
    
    [[did]] = call_procedure('GetDepartmentIDFromNameSite', (name, site))
    return jsonify({'did': did})

@app.route('/api/generate-excel', methods=['GET'])
def generate_excel():
    data_param = request.args.get('data')
    data = json.loads(data_param)

    df = pd.DataFrame(data)

    excel_writer = pd.ExcelWriter('downloads/attendance_reports/AttendanceReport.xlsx', engine='openpyxl')
    excel_writer.book = Workbook()

    df.to_excel(excel_writer, index=False, sheet_name='Sheet')

    sheet = excel_writer.sheets['Sheet']
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    excel_writer.save()
    
    return send_file('downloads/attendance_reports/AttendanceReport.xlsx', as_attachment=True)

# ---------------------------------------

if __name__ == '__main__':
    app.run(debug=True, port=5000)
